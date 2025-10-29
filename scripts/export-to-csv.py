#!/usr/bin/env python3
"""
Export each sheet from Daily Sales Report.xlsx to individual CSV files
"""

import openpyxl
import csv
import os

# Load the workbook
print("📂 Loading Daily Sales Report.xlsx...")
wb = openpyxl.load_workbook('Daily Sales Report.xlsx', data_only=True)

# Create output directory
output_dir = 'csv-exports'
os.makedirs(output_dir, exist_ok=True)

print(f"📊 Found {len(wb.sheetnames)} sheets")
print(f"📁 Exporting to {output_dir}/\n")

# Skip first sheet (summary), export the rest
sheets_exported = 0
for i, sheet_name in enumerate(wb.sheetnames[1:], start=2):  # Start from sheet 2
    sheet = wb[sheet_name]

    # Create CSV filename (sanitize sheet name)
    csv_filename = f"{output_dir}/{sheet_name.replace('/', '-')}.csv"

    # Write to CSV
    with open(csv_filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        for row in sheet.iter_rows(values_only=True):
            # Convert None to empty string
            cleaned_row = ['' if cell is None else cell for cell in row]
            writer.writerow(cleaned_row)

    sheets_exported += 1
    print(f"✅ Exported: {sheet_name} → {csv_filename}")

print(f"\n✅ Done! Exported {sheets_exported} sheets to CSV")
